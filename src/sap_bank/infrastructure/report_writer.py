"""
report_writer.py
----------------
Genera el reporte de conciliacion .xlsx por cada CSV procesado, con relleno
de color por estado (verde=EXITO, rojo=ERROR, amarillo=OBSERVADO).
"""

from __future__ import annotations

import logging
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.sap_bank.domain.models import FileSummary, RowStatus

logger = logging.getLogger(__name__)

HEADERS = ["N Linea", "Estado", "DocEntry SAP", "DocNum SAP", "Cuenta Destino", "Error"]

FILL_BY_STATUS = {
    RowStatus.EXITO: PatternFill("solid", fgColor="C6EFCE"),
    RowStatus.ERROR: PatternFill("solid", fgColor="FFC7CE"),
    RowStatus.OBSERVADO: PatternFill("solid", fgColor="FFEB9C"),
    RowStatus.OMITIDA: PatternFill("solid", fgColor="DDEBF7"),
}
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_report(summary: FileSummary, reports_dir: str) -> str:
    """Escribe el .xlsx y devuelve la ruta del archivo generado."""
    os.makedirs(reports_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(summary.source_name)[0]
    out_path = os.path.join(reports_dir, f"REPORTE_{base}_{ts}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliacion"

    # Encabezado
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Filas
    for i, r in enumerate(summary.results, start=2):
        values = [r.line_num, r.status.value, r.doc_entry, r.doc_num, r.cuenta_destino, r.error]
        fill = FILL_BY_STATUS.get(r.status)
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    _autofit(ws)
    ws.freeze_panes = "A2"

    # Hoja resumen
    _write_summary_sheet(wb, summary)

    wb.save(out_path)
    logger.info("Reporte generado: %s", out_path)
    return out_path


def _write_summary_sheet(wb: Workbook, summary: FileSummary) -> None:
    ws = wb.create_sheet("Resumen")
    rows = [
        ("Archivo origen", summary.source_name),
        ("Fecha de proceso", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total filas", summary.total),
        ("EXITO", summary.exitos),
        ("ERROR", summary.errores),
        ("OBSERVADO", summary.observados),
        ("OMITIDA (ya existia)", summary.omitidas),
        ("Resultado", "OK" if summary.all_success else "CON INCIDENCIAS"),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws)


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(width + 2, 10), 60)
